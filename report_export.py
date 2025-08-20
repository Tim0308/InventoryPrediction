import os
from typing import List, Tuple
import pandas as pd
from openpyxl.styles import PatternFill


def create_excel_from_prediction_summary(
    hospital_dfs: List[Tuple[str, pd.DataFrame]],
    output_path: str = "predictions_summary.xlsx",
) -> str:
    """
    Create an Excel report from a list of (hospital_name, results_df) tuples.

    Each results_df is expected to have columns: 'Date', 'Time After Last Invoices', 'Type'.
    One row should have Type == 'Forecasted Next Invoice Date (Stacking)' for the next forecast.

    Returns the path to the created Excel file.
    """
    if not hospital_dfs:
        raise ValueError("No prediction summaries provided.")

    # Build summary rows
    summary_rows = []
    for hospital_name, product_name, df in hospital_dfs:
        if df is None or df.empty:
            continue
        df = df.copy()
        # Ensure datetime
        if not pd.api.types.is_datetime64_any_dtype(df['Date']):
            df['Date'] = pd.to_datetime(df['Date'])

        # Forecast info T+1..T+4 (prefer explicit T+K labels if available)
        def pick_forecast(k: int):
            label_explicit = f'Forecasted Next Invoice Date T+{k} (Stacking)'
            row = df[df['Type'] == label_explicit]
            if not row.empty:
                return row['Date'].iloc[-1], row['Time After Last Invoices'].iloc[-1]
            # Fallback to legacy single-step label for T+1
            if k == 1:
                legacy = df[df['Type'] == 'Forecasted Next Invoice Date (Stacking)']
                if not legacy.empty:
                    return legacy['Date'].iloc[-1], legacy['Time After Last Invoices'].iloc[-1]
            return pd.NaT, None

        f_dates = {}
        f_intervals = {}
        for k in (1, 2, 3, 4):
            d, i = pick_forecast(k)
            f_dates[k] = d
            f_intervals[k] = i

        # Last actual date (based on the max date among actual/predicted series)
        historical_mask = df['Type'].isin(['Actual', 'Predicted (Stacking)'])
        last_actual_date = df.loc[historical_mask, 'Date'].max() if historical_mask.any() else df['Date'].max()

        def fmt_date(x):
            return x.strftime('%Y-%m-%d') if pd.notna(x) else None
        def fmt_int(x):
            return int(x) if pd.notna(pd.Series([x])[0]) else None

        summary_rows.append({
            'Product': product_name,
            'Hospital': hospital_name,
            'LastInvoiceDate': fmt_date(last_actual_date),
            # 'PredictedIntervalDays_T+1': fmt_int(f_intervals[1]),
            'ForecastedDate_T+1': fmt_date(f_dates[1]),
            # 'PredictedIntervalDays_T+2': fmt_int(f_intervals[2]),
            'ForecastedDate_T+2': fmt_date(f_dates[2]),
            # 'PredictedIntervalDays_T+3': fmt_int(f_intervals[3]),
            'ForecastedDate_T+3': fmt_date(f_dates[3]),
            # 'PredictedIntervalDays_T+4': fmt_int(f_intervals[4]),
            'ForecastedDate_T+4': fmt_date(f_dates[4]),
        })

    summary_df = pd.DataFrame(summary_rows)

    # Write Excel with summary + per-hospital sheets
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        # Per-hospital sheets (trim to reasonable length if very large)
        for hospital_name, product_name, df in hospital_dfs:
            if df is None or df.empty:
                continue
            df_export = df.copy()
            # Rename for human readability
            df_export = df_export.rename(columns={
                'Date': 'InvoiceDate',
                'Time After Last Invoices': 'Time_Between_Invoices'
            })
            # Sort and write
            df_export = df_export.sort_values('InvoiceDate')
            safe_sheet = hospital_name
            # Excel sheet name limit 31 chars; ensure uniqueness by truncation
            if len(safe_sheet) > 31:
                safe_sheet = safe_sheet[:31]
            df_export.to_excel(writer, sheet_name=safe_sheet, index=False)

        # Highlight ForecastedDate_T+1 column in Summary sheet (yellow fill)
        try:
            ws = writer.sheets.get('Summary')
            if ws is not None and 'ForecastedDate_T+1' in summary_df.columns:
                col_idx = summary_df.columns.get_loc('ForecastedDate_T+1') + 1  # 1-based for openpyxl
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                # Apply fill to all data rows (skip header at row 1)
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = yellow_fill
        except Exception:
            # Do not fail report generation if styling fails
            pass

    return output_path
